import asyncio
import json
import os
import re
import shutil
from io import BytesIO
from pathlib import Path
from datetime import datetime, timedelta, date as date_type
from fastapi import FastAPI, HTTPException, Depends, Request, UploadFile, File, Form, Header
from fastapi.middleware.cors import CORSMiddleware
from fastapi.staticfiles import StaticFiles
from fastapi.responses import FileResponse, StreamingResponse
from sqlalchemy.orm import Session, sessionmaker
from sqlalchemy import func, text
from pydantic import BaseModel
from typing import Optional
from openpyxl import Workbook, load_workbook
from jose import JWTError, jwt

from models import ENGINE, init_db, Product, Order, InventoryItem, InventoryType, MappingRule

# ── JWT 설정 ──────────────────────────────────────────────
_JWT_SECRET  = os.getenv('JWT_SECRET', 'yak-jwt-secret-2026')
_JWT_ALG     = 'HS256'
_TOKEN_EXP_H = 24  # 토큰 유효시간 (시간)
_ADMIN_PW    = os.getenv('ADMIN_PASSWORD',  'newface')
_VIEWER_PW   = os.getenv('VIEWER_PASSWORD', 'blackyak')
_BACKUP_TOKEN = os.getenv('BACKUP_TOKEN', '')  # Cloud Scheduler 전용 토큰

def _create_token(role: str) -> str:
    exp = datetime.utcnow() + timedelta(hours=_TOKEN_EXP_H)
    return jwt.encode({'sub': role, 'exp': exp}, _JWT_SECRET, algorithm=_JWT_ALG)

def _verify_token(authorization: Optional[str] = Header(None)) -> str:
    if not authorization or not authorization.startswith('Bearer '):
        raise HTTPException(401, '인증이 필요합니다')
    try:
        payload = jwt.decode(authorization.split(' ')[1], _JWT_SECRET, algorithms=[_JWT_ALG])
        return payload.get('sub', '')
    except JWTError:
        raise HTTPException(401, '유효하지 않은 토큰입니다')

def _require_admin(role: str = Depends(_verify_token)) -> str:
    if role != 'admin':
        raise HTTPException(403, '관리자 권한이 필요합니다')
    return role

init_db()
with ENGINE.connect() as _conn:
    try:
        _conn.execute(text('ALTER TABLE products ADD COLUMN active INTEGER DEFAULT 1'))
        _conn.commit()
    except Exception:
        pass  # column already exists
SessionLocal = sessionmaker(bind=ENGINE)

# ── 날짜 자동 마이그레이션: M.DD → YYYY-MM-DD ────────────
_DATE_OLD = re.compile(r'^(\d{1,2})\.(\d{2})$')

def _migrate_dates():
    """기존 M.DD 형식 날짜를 YYYY-MM-DD 로 일괄 변환 (앱 시작 시 1회)."""
    db = SessionLocal()
    try:
        year = datetime.now().year
        changed = 0
        for model in (Order, InventoryItem):
            for row in db.query(model).all():
                m = _DATE_OLD.match(row.date or '')
                if m:
                    row.date = f'{year}-{int(m.group(1)):02d}-{int(m.group(2)):02d}'
                    changed += 1
        if changed:
            db.commit()
            _sync_db()
            print(f'[migrate] 날짜 {changed}건 변환 완료 → YYYY-MM-DD')
    except Exception as e:
        print(f'[migrate] 실패: {e}')
    finally:
        db.close()

# ── GCS 동기화: 쓰기 후 /tmp/yak.db → /data/yak.db 복사 ──
def _sync_db():
    """쓰기 작업 완료 후 /tmp DB를 GCS FUSE(/data)로 동기화."""
    if os.path.exists('/tmp/yak.db') and os.path.isdir('/data'):
        try:
            shutil.copy2('/tmp/yak.db', '/data/yak.db')
        except Exception as e:
            print(f'[sync] {e}')

# ── 시드 데이터 (DB가 비어있을 때 상품 자동 등록) ─────────
def _do_seed(db):
    _seed_file = Path(__file__).parent / 'seed_products.json'
    if not _seed_file.exists():
        return 0
    data = json.loads(_seed_file.read_text(encoding='utf-8'))
    for p in data:
        db.add(Product(
            name=p['name'], color=p['color'],
            size=p['size'], model_code=p.get('model_code')
        ))
    db.commit()
    return len(data)

def _seed_products():
    try:
        db = SessionLocal()
        try:
            if db.query(Product).count() == 0:
                n = _do_seed(db)
                print(f'[seed] 상품 {n}개 등록 완료')
                _sync_db()  # 최초 시드 후 GCS에 반영
        finally:
            db.close()
    except Exception as e:
        print(f'[seed] 실패: {e}')

_seed_products()
_migrate_dates()

app = FastAPI(title='야크 재고관리 API')

# ── GCS 동기화 미들웨어: POST/PUT/DELETE 완료 후 동기 sync ──
@app.middleware('http')
async def _gcs_sync_middleware(request: Request, call_next):
    response = await call_next(request)
    if request.method in ('POST', 'PUT', 'DELETE', 'PATCH'):
        await asyncio.to_thread(_sync_db)
    return response

app.add_middleware(
    CORSMiddleware,
    allow_origins=['*'],
    allow_methods=['*'],
    allow_headers=['*'],
)


def get_db():
    db = SessionLocal()
    try:
        yield db
    finally:
        db.close()


# ─── Pydantic 스키마 ──────────────────────────────────────

class ProductIn(BaseModel):
    name: str
    color: str
    size: str
    model_code: str = ''

class ProductOut(BaseModel):
    id: int
    name: str
    color: str
    size: str
    model_code: str
    active: bool = True
    model_config = {'from_attributes': True}

class OrderIn(BaseModel):
    date: str
    product_id: int
    quantity: int
    order_date: str = ''
    storage: str = '뉴페이스'
    mall: str = ''
    orderer: str = ''
    recipient: str = ''
    phone: str = ''
    address: str = ''
    memo: str = ''

class OrderOut(OrderIn):
    id: int
    product: Optional[ProductOut] = None
    created_at: datetime
    model_config = {'from_attributes': True}

class InventoryIn(BaseModel):
    date: str
    product_id: int
    quantity: int
    type: InventoryType = InventoryType.normal
    notes: str = ''

class InventoryOut(InventoryIn):
    id: int
    product: Optional[ProductOut] = None
    created_at: datetime
    model_config = {'from_attributes': True}

class StockSummaryOut(BaseModel):
    product: ProductOut
    total_in: int
    total_out: int
    current_stock: int
    low_stock: bool

class DailyOutboundOut(BaseModel):
    date: str
    product_id: int
    quantity: int

class MappingRuleIn(BaseModel):
    rule_name: str = ''
    product_id: Optional[int] = None
    match_type: str = 'and'    # 'and' | 'or'
    keywords: list[str] = []
    enabled: bool = True
    priority: int = 0

class MappingRuleOut(MappingRuleIn):
    id: int
    product: Optional[ProductOut] = None
    created_at: datetime
    model_config = {'from_attributes': True}

    @classmethod
    def from_orm_rule(cls, rule: MappingRule) -> 'MappingRuleOut':
        return cls(
            id=rule.id,
            rule_name=rule.rule_name,
            product_id=rule.product_id,
            match_type=rule.match_type,
            keywords=json.loads(rule.keywords or '[]'),
            enabled=rule.enabled,
            priority=rule.priority,
            created_at=rule.created_at,
            product=rule.product,
        )

class ResolveIn(BaseModel):
    product_name: str   # 상품명 원본 텍스트

class LoginIn(BaseModel):
    password: str


# ─── 인증 ────────────────────────────────────────────────

@app.post('/auth/login')
def login(data: LoginIn):
    if data.password == _ADMIN_PW:
        return {'token': _create_token('admin'), 'role': 'admin'}
    if data.password == _VIEWER_PW:
        return {'token': _create_token('viewer'), 'role': 'viewer'}
    raise HTTPException(401, '비밀번호가 틀렸습니다')


@app.post('/admin/seed-products')
def seed_products_api(db: Session = Depends(get_db), _: str = Depends(_require_admin)):
    cnt = db.query(Product).count()
    if cnt > 0:
        return {'ok': False, 'message': f'이미 {cnt}개 상품 존재'}
    n = _do_seed(db)
    return {'ok': True, 'inserted': n}

# ─── 제품 ────────────────────────────────────────────────

@app.get('/products', response_model=list[ProductOut])
def get_products(db: Session = Depends(get_db)):
    return db.query(Product).order_by(Product.name, Product.color, Product.size).all()


@app.post('/products', response_model=ProductOut)
def create_product(data: ProductIn, db: Session = Depends(get_db), _: str = Depends(_require_admin)):
    # 중복 체크
    exists = db.query(Product).filter(
        Product.name == data.name,
        Product.color == data.color,
        Product.size == data.size,
    ).first()
    if exists:
        raise HTTPException(400, '이미 존재하는 제품입니다')
    p = Product(**data.model_dump())
    db.add(p)
    db.commit()
    db.refresh(p)
    return p


@app.put('/products/{product_id}', response_model=ProductOut)
def update_product(product_id: int, data: ProductIn, db: Session = Depends(get_db), _: str = Depends(_require_admin)):
    p = db.get(Product, product_id)
    if not p:
        raise HTTPException(404, '제품을 찾을 수 없습니다')
    p.name = data.name
    p.color = data.color
    p.size = data.size
    p.model_code = data.model_code
    db.commit()
    db.refresh(p)
    return p


@app.delete('/products/{product_id}')
def delete_product(product_id: int, db: Session = Depends(get_db), _: str = Depends(_require_admin)):
    p = db.get(Product, product_id)
    if not p:
        raise HTTPException(404, '제품을 찾을 수 없습니다')
    db.delete(p)
    db.commit()
    return {'ok': True}


@app.patch('/products/{product_id}/toggle-active', response_model=ProductOut)
def toggle_product_active(product_id: int, db: Session = Depends(get_db), _: str = Depends(_require_admin)):
    p = db.get(Product, product_id)
    if not p:
        raise HTTPException(404, '제품을 찾을 수 없습니다')
    p.active = not (p.active if p.active is not None else True)
    db.commit()
    db.refresh(p)
    return p


# ─── 발주 ────────────────────────────────────────────────

@app.get('/orders', response_model=list[OrderOut])
def get_orders(
    month: Optional[str] = None,
    date: Optional[str] = None,
    db: Session = Depends(get_db),
    _role: str = Depends(_verify_token),
):
    q = db.query(Order)
    if month:
        # month: "2026-04" → LIKE '2026-04-%'
        q = q.filter(Order.date.like(f'{month}-%'))
    if date:
        q = q.filter(Order.date == date)
    return q.order_by(Order.date).all()


@app.post('/orders', response_model=OrderOut)
def create_order(data: OrderIn, db: Session = Depends(get_db), _: str = Depends(_require_admin)):
    if not db.get(Product, data.product_id):
        raise HTTPException(404, '제품을 찾을 수 없습니다')
    order = Order(**data.model_dump())
    db.add(order)
    db.commit()
    db.refresh(order)
    return order


@app.put('/orders/{order_id}', response_model=OrderOut)
def update_order(order_id: int, data: OrderIn, db: Session = Depends(get_db), _: str = Depends(_require_admin)):
    order = db.get(Order, order_id)
    if not order:
        raise HTTPException(404, '발주를 찾을 수 없습니다')
    if not db.get(Product, data.product_id):
        raise HTTPException(404, '제품을 찾을 수 없습니다')
    for k, v in data.model_dump().items():
        setattr(order, k, v)
    db.commit()
    db.refresh(order)
    return order


@app.delete('/orders/{order_id}')
def delete_order(order_id: int, db: Session = Depends(get_db), _: str = Depends(_require_admin)):
    order = db.get(Order, order_id)
    if not order:
        raise HTTPException(404, '발주를 찾을 수 없습니다')
    db.delete(order)
    db.commit()
    return {'ok': True}

@app.post('/orders/batch-delete')
def batch_delete_orders(body: dict, db: Session = Depends(get_db), _: str = Depends(_require_admin)):
    from sqlalchemy import delete as sa_delete
    ids: list[int] = body.get('ids', [])
    if not ids:
        return {'deleted': 0}
    result = db.execute(sa_delete(Order).where(Order.id.in_(ids)))
    db.commit()
    return {'deleted': result.rowcount}

@app.post('/inventory/batch-delete')
def batch_delete_inventory(body: dict, db: Session = Depends(get_db), _: str = Depends(_require_admin)):
    from sqlalchemy import delete as sa_delete
    ids: list[int] = body.get('ids', [])
    if not ids:
        return {'deleted': 0}
    result = db.execute(sa_delete(InventoryItem).where(InventoryItem.id.in_(ids)))
    db.commit()
    return {'deleted': result.rowcount}

@app.post('/orders/bulk')
def create_orders_bulk(data: list[OrderIn], db: Session = Depends(get_db), _: str = Depends(_require_admin)):
    if len(data) > 500:
        raise HTTPException(400, f'한 번에 최대 500건까지 등록 가능합니다 (요청: {len(data)}건)')
    for order_data in data:
        db.add(Order(**order_data.model_dump()))
    db.commit()
    return {'ok': len(data), 'fail': []}


# ─── 입고 ────────────────────────────────────────────────

@app.get('/inventory', response_model=list[InventoryOut])
def get_inventory(month: Optional[str] = None, db: Session = Depends(get_db), _role: str = Depends(_verify_token)):
    q = db.query(InventoryItem)
    if month:
        q = q.filter(InventoryItem.date.like(f'{month}-%'))
    return q.order_by(InventoryItem.date).all()


@app.post('/inventory', response_model=InventoryOut)
def create_inventory(data: InventoryIn, db: Session = Depends(get_db), _: str = Depends(_require_admin)):
    if not db.get(Product, data.product_id):
        raise HTTPException(404, '제품을 찾을 수 없습니다')
    item = InventoryItem(**data.model_dump())
    db.add(item)
    db.commit()
    db.refresh(item)
    return item


@app.post('/inventory/bulk')
def create_inventory_bulk(data: list[InventoryIn], db: Session = Depends(get_db), _: str = Depends(_require_admin)):
    if len(data) > 500:
        raise HTTPException(400, f'한 번에 최대 500건까지 등록 가능합니다 (요청: {len(data)}건)')
    for item_data in data:
        db.add(InventoryItem(**item_data.model_dump()))
    db.commit()
    return {'ok': len(data), 'fail': []}


@app.put('/inventory/{item_id}', response_model=InventoryOut)
def update_inventory(item_id: int, data: InventoryIn, db: Session = Depends(get_db), _: str = Depends(_require_admin)):
    item = db.get(InventoryItem, item_id)
    if not item:
        raise HTTPException(404, '입고 항목을 찾을 수 없습니다')
    if not db.get(Product, data.product_id):
        raise HTTPException(404, '제품을 찾을 수 없습니다')
    for k, v in data.model_dump().items():
        setattr(item, k, v)
    db.commit()
    db.refresh(item)
    return item


@app.delete('/inventory/{item_id}')
def delete_inventory(item_id: int, db: Session = Depends(get_db), _: str = Depends(_require_admin)):
    item = db.get(InventoryItem, item_id)
    if not item:
        raise HTTPException(404, '입고 항목을 찾을 수 없습니다')
    db.delete(item)
    db.commit()
    return {'ok': True}


# ─── 재고 현황 ───────────────────────────────────────────

@app.get('/stock/summary', response_model=list[StockSummaryOut])
def get_stock_summary(month: Optional[str] = None, db: Session = Depends(get_db), _role: str = Depends(_verify_token)):
    products = db.query(Product).order_by(Product.name, Product.color, Product.size).all()

    inv_q = db.query(InventoryItem.product_id, func.sum(InventoryItem.quantity).label('total'))
    if month:
        inv_q = inv_q.filter(InventoryItem.date.like(f'{month}-%'))
    # 불량 입고는 현재고에서 제외
    inv_q = inv_q.filter(InventoryItem.type != InventoryType.defective)
    inv_map = {r.product_id: r.total for r in inv_q.group_by(InventoryItem.product_id).all()}

    ord_q = db.query(Order.product_id, func.sum(Order.quantity).label('total'))
    if month:
        ord_q = ord_q.filter(Order.date.like(f'{month}-%'))
    ord_map = {r.product_id: r.total for r in ord_q.group_by(Order.product_id).all()}

    result = []
    for p in products:
        total_in  = inv_map.get(p.id, 0)
        total_out = ord_map.get(p.id, 0)
        current   = total_in - total_out
        result.append(StockSummaryOut(
            product=p, total_in=total_in, total_out=total_out,
            current_stock=current, low_stock=current < 1,
        ))
    return result


@app.get('/stock/daily', response_model=list[DailyOutboundOut])
def get_daily_outbound(month: str, db: Session = Depends(get_db), _role: str = Depends(_verify_token)):
    rows = db.query(
        Order.date, Order.product_id,
        func.sum(Order.quantity).label('quantity'),
    ).filter(Order.date.like(f'{month}-%')).group_by(Order.date, Order.product_id).all()
    return [DailyOutboundOut(date=r.date, product_id=r.product_id, quantity=r.quantity)
            for r in rows]


# ─── 매핑 규칙 ───────────────────────────────────────────

@app.get('/mapping-rules', response_model=list[MappingRuleOut])
def get_mapping_rules(db: Session = Depends(get_db)):
    rules = db.query(MappingRule).order_by(MappingRule.priority.desc(), MappingRule.id).all()
    return [MappingRuleOut.from_orm_rule(r) for r in rules]


@app.post('/mapping-rules', response_model=MappingRuleOut)
def create_mapping_rule(data: MappingRuleIn, db: Session = Depends(get_db), _: str = Depends(_require_admin)):
    rule = MappingRule(
        rule_name=data.rule_name,
        product_id=data.product_id,
        match_type=data.match_type,
        keywords=json.dumps(data.keywords, ensure_ascii=False),
        enabled=data.enabled,
        priority=data.priority,
    )
    db.add(rule)
    db.commit()
    db.refresh(rule)
    return MappingRuleOut.from_orm_rule(rule)


@app.put('/mapping-rules/{rule_id}', response_model=MappingRuleOut)
def update_mapping_rule(rule_id: int, data: MappingRuleIn, db: Session = Depends(get_db), _: str = Depends(_require_admin)):
    rule = db.get(MappingRule, rule_id)
    if not rule:
        raise HTTPException(404, '규칙을 찾을 수 없습니다')
    rule.rule_name  = data.rule_name
    rule.product_id = data.product_id
    rule.match_type = data.match_type
    rule.keywords   = json.dumps(data.keywords, ensure_ascii=False)
    rule.enabled    = data.enabled
    rule.priority   = data.priority
    db.commit()
    db.refresh(rule)
    return MappingRuleOut.from_orm_rule(rule)


@app.delete('/mapping-rules/{rule_id}')
def delete_mapping_rule(rule_id: int, db: Session = Depends(get_db), _: str = Depends(_require_admin)):
    rule = db.get(MappingRule, rule_id)
    if not rule:
        raise HTTPException(404, '규칙을 찾을 수 없습니다')
    db.delete(rule)
    db.commit()
    return {'ok': True}


@app.patch('/mapping-rules/{rule_id}/toggle')
def toggle_mapping_rule(rule_id: int, db: Session = Depends(get_db), _: str = Depends(_require_admin)):
    rule = db.get(MappingRule, rule_id)
    if not rule:
        raise HTTPException(404, '규칙을 찾을 수 없습니다')
    rule.enabled = not rule.enabled
    db.commit()
    return {'id': rule.id, 'enabled': rule.enabled}


@app.post('/mapping-rules/resolve')
def resolve_product(data: ResolveIn, db: Session = Depends(get_db)):
    """
    상품명 텍스트 → product_id 자동 해석.
    우선순위 높은 규칙부터 순서대로 검사.
    """
    text = data.product_name.strip()
    rules = db.query(MappingRule).filter(MappingRule.enabled == True).order_by(
        MappingRule.priority.desc(), MappingRule.id
    ).all()

    for rule in rules:
        keywords = json.loads(rule.keywords or '[]')
        if not keywords:
            continue
        normalized = text.upper()
        kws = [k.upper() for k in keywords]
        if rule.match_type == 'and':
            matched = all(k in normalized for k in kws)
        else:  # 'or'
            matched = any(k in normalized for k in kws)
        if matched and rule.product_id:
            product = db.get(Product, rule.product_id)
            if product:
                return {'product_id': rule.product_id, 'product': ProductOut.model_validate(product)}
    return {'product_id': None, 'product': None}


@app.post('/mapping-rules/seed-defaults')
def seed_default_rules(db: Session = Depends(get_db), _: str = Depends(_require_admin)):
    """
    스프레드시트 출고관리 검색조건 기반 기본 규칙 자동 생성.
    이미 규칙이 있으면 건너뜀.
    """
    cnt = db.query(MappingRule).count()
    if cnt > 0:
        return {'message': '이미 규칙이 존재합니다', 'count': cnt}

    products = {(p.name, p.color, p.size): p.id
                for p in db.query(Product).all()}

    rules_to_create = []

    # OR 방식 - 용품 3종
    bag_rules = [
        ('야크커뮤트 힙색',   '블랙', '단품', ['힙색', '8BYABF3902'],  'or'),
        ('야크커뮤트 슬링백', '블랙', '단품', ['슬링백', '8BYABF3901'], 'or'),
        ('야크커뮤트 백팩',   '블랙', '단품', ['백팩', '8BYKSF3901'],  'or'),
    ]
    for name, color, size, kws, mtype in bag_rules:
        pid = products.get((name, color, size))
        if pid:
            rules_to_create.append(MappingRule(
                rule_name=f'{name}',
                product_id=pid,
                match_type=mtype,
                keywords=json.dumps(kws, ensure_ascii=False),
                enabled=True, priority=10,
            ))

    # AND 방식 - 의류 (제품명 키워드 + 색상 + 사이즈)
    clothing = [
        ('H티아고 자켓',    '블랙',       ['티아고', '블랙']),
        ('H티아고 자켓',    '그레이',     ['티아고', '그레이']),
        ('H포그 윈드자켓',  '그레이',     ['윈드', '그레이']),
        ('H포그 윈드자켓',  '카키',       ['윈드', '카키']),
        ('H미토 윈드자켓',  '블랙',       ['미토', '블랙']),
        ('H주빌로 자켓',    '블루',       ['주빌로', '블루']),
        ('H주빌로 자켓',    '라이트베이지', ['주빌로', '베이지']),
        ('H주빌로 자켓',    '라이트그레이', ['주빌로', '그레이']),
        ('H피레스코 티셔츠', '블랙',      ['피레스코', '블랙']),
        ('H피레스코 티셔츠', '화이트',    ['피레스코', '화이트']),
    ]
    sizes = ['90', '95', '100', '105', '110', '115']
    for name, color, base_kws in clothing:
        for size in sizes:
            pid = products.get((name, color, size))
            if not pid:
                continue
            kws = base_kws + [size]
            rules_to_create.append(MappingRule(
                rule_name=f'{name} / {color} / {size}',
                product_id=pid,
                match_type='and',
                keywords=json.dumps(kws, ensure_ascii=False),
                enabled=True, priority=5,
            ))

    # 미토 윈드자켓 단품 사이즈 (95~115)
    mito_sizes = ['95', '100', '105', '110', '115']
    for size in mito_sizes:
        pid = products.get(('H미토 윈드자켓', '블랙', size))
        if pid:
            rules_to_create.append(MappingRule(
                rule_name=f'H미토 윈드자켓 / 블랙 / {size}',
                product_id=pid,
                match_type='and',
                keywords=json.dumps(['미토', '블랙', size], ensure_ascii=False),
                enabled=True, priority=5,
            ))

    for rule in rules_to_create:
        db.add(rule)
    db.commit()
    return {'message': f'{len(rules_to_create)}개 기본 규칙 생성 완료', 'count': len(rules_to_create)}


# ─── 백업 / 복원 ─────────────────────────────────────────────

@app.post('/backup/auto')
def backup_auto(x_backup_token: Optional[str] = Header(None)):
    """
    Cloud Scheduler가 매일 새벽 호출 → /data/backup/ 폴더에 날짜별 DB 스냅샷 저장.
    BACKUP_TOKEN 환경변수로 인증 (일반 JWT와 별도).
    X-Backup-Token 헤더 사용 (Authorization 헤더는 Cloud Scheduler가 override함).
    """
    # 토큰 검증
    if not _BACKUP_TOKEN:
        raise HTTPException(503, 'BACKUP_TOKEN 환경변수가 설정되지 않았습니다')
    if not x_backup_token or x_backup_token != _BACKUP_TOKEN:
        raise HTTPException(401, '백업 토큰이 유효하지 않습니다')

    # 실제 DB 경로 (SQLAlchemy ENGINE URL에서 추출)
    src = Path(ENGINE.url.database)
    if not src.exists():
        raise HTTPException(503, f'{src} 파일이 없습니다')

    backup_dir = Path('/data/backup')
    try:
        backup_dir.mkdir(parents=True, exist_ok=True)
        stamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        dst = backup_dir / f'yak_{stamp}.db'
        shutil.copy2(str(src), str(dst))

        # 30일 이상 된 백업 자동 삭제
        cutoff = datetime.now().timestamp() - 30 * 86400
        removed = 0
        for old in backup_dir.glob('yak_*.db'):
            if old.stat().st_mtime < cutoff:
                old.unlink()
                removed += 1

        return {'ok': True, 'saved': str(dst), 'removed_old': removed}
    except Exception as e:
        raise HTTPException(500, f'백업 실패: {e}')


@app.get('/backup/export')
def backup_export(db: Session = Depends(get_db)):
    wb = Workbook()

    ws_p = wb.active
    ws_p.title = 'products'
    ws_p.append(['id', 'name', 'color', 'size', 'model_code'])
    for p in db.query(Product).order_by(Product.id).all():
        ws_p.append([p.id, p.name, p.color, p.size, p.model_code or ''])

    ws_o = wb.create_sheet('orders')
    ws_o.append(['id', 'date', 'product_id', 'quantity', 'order_date',
                 'storage', 'mall', 'orderer', 'recipient', 'phone', 'address', 'memo'])
    for o in db.query(Order).order_by(Order.date, Order.id).all():
        ws_o.append([o.id, o.date, o.product_id, o.quantity,
                     o.order_date or '', o.storage or '', o.mall or '',
                     o.orderer or '', o.recipient or '', o.phone or '',
                     o.address or '', o.memo or ''])

    ws_i = wb.create_sheet('inventory')
    ws_i.append(['id', 'date', 'product_id', 'quantity', 'type', 'notes'])
    for it in db.query(InventoryItem).order_by(InventoryItem.date, InventoryItem.id).all():
        ws_i.append([it.id, it.date, it.product_id, it.quantity,
                     it.type.value, it.notes or ''])

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f'yak_backup_{date_type.today().strftime("%Y%m%d")}.xlsx'
    return StreamingResponse(
        buf,
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f'attachment; filename="{filename}"'},
    )


@app.post('/backup/import')
async def backup_import(
    file: UploadFile = File(...),
    mode: str = Form('append'),
    db: Session = Depends(get_db),
    _: str = Depends(_require_admin),
):
    content = await file.read()
    try:
        wb = load_workbook(BytesIO(content), data_only=True)
    except Exception:
        raise HTTPException(400, '올바른 xlsx 파일이 아닙니다')

    if mode == 'reset':
        db.query(Order).delete()
        db.query(InventoryItem).delete()
        db.commit()

    stats = {'orders': 0, 'inventory': 0}

    if 'orders' in wb.sheetnames:
        rows = list(wb['orders'].iter_rows(values_only=True))
        for row in rows[1:]:
            if not row or row[0] is None:
                continue
            try:
                row = tuple(row) + (None,) * 12
                _, dt, pid, qty, odt, storage, mall, orderer, recipient, phone, address, memo = row[:12]
                if not dt or not pid or qty is None:
                    continue
                db.add(Order(
                    date=str(dt), product_id=int(pid), quantity=int(qty),
                    order_date=str(odt or ''), storage=str(storage or '뉴페이스'),
                    mall=str(mall or ''), orderer=str(orderer or ''),
                    recipient=str(recipient or ''), phone=str(phone or ''),
                    address=str(address or ''), memo=str(memo or ''),
                ))
                stats['orders'] += 1
            except Exception:
                continue

    if 'inventory' in wb.sheetnames:
        rows = list(wb['inventory'].iter_rows(values_only=True))
        for row in rows[1:]:
            if not row or row[0] is None:
                continue
            try:
                row = tuple(row) + (None,) * 6
                _, dt, pid, qty, type_val, notes = row[:6]
                if not dt or not pid or qty is None:
                    continue
                try:
                    inv_type = InventoryType(str(type_val or 'normal'))
                except ValueError:
                    inv_type = InventoryType.normal
                db.add(InventoryItem(
                    date=str(dt), product_id=int(pid), quantity=int(qty),
                    type=inv_type, notes=str(notes or ''),
                ))
                stats['inventory'] += 1
            except Exception:
                continue

    db.commit()
    return stats


# ─── React SPA 서빙 (프로덕션 빌드) ──────────────────────────
# 반드시 모든 API 라우트 등록 후 마지막에 위치
_static = Path(__file__).parent / 'static'
if _static.exists():
    # /assets, /vite.svg 등 정적 파일
    app.mount('/assets', StaticFiles(directory=str(_static / 'assets')), name='spa-assets')

    @app.get('/{full_path:path}', include_in_schema=False)
    async def serve_spa(full_path: str):
        """SPA fallback — React Router가 처리하도록 index.html 반환"""
        file_path = _static / full_path
        if file_path.exists() and file_path.is_file():
            return FileResponse(str(file_path))
        return FileResponse(str(_static / 'index.html'))
